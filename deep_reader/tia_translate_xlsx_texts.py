from __future__ import annotations

import argparse
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

import openpyxl


_CJK_RE = re.compile(r"[\u4e00-\u9fff]")
_LATIN_RE = re.compile(r"[A-Za-z]")

# Keep placeholders/tokens intact for WinCC/TIA import compatibility.
_PLACEHOLDER_RE = re.compile(
    r"(\r\n|\n)"  # line breaks
    r"|(\{[^}]+\})"  # {0}, {Tag}, etc.
    r"|(%\d*\$?[sdif])"  # %s, %1$d, etc.
    r"|(@[A-Za-z0-9_.]+)"  # @Tag
    r"|(\$[A-Za-z0-9_.]+)"  # $Tag
    r"|(<[^>]+>)"  # simple markup like <br/>
)


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _classify_text(text: str) -> str:
    if not text:
        return "empty"
    if _CJK_RE.search(text):
        return "cn"
    if _LATIN_RE.search(text):
        return "latin"
    return "other"


@dataclass(frozen=True)
class Columns:
    zh_star: int
    zh: int
    en: int
    sk: int


def _find_columns(headers: Sequence[object]) -> Columns:
    idx = {str(v): i for i, v in enumerate(headers) if v is not None}
    required = ["zh-CN*", "zh-CN", "en-GB", "sk-SK"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise KeyError(f"Missing required columns: {missing}")
    return Columns(
        zh_star=idx["zh-CN*"],
        zh=idx["zh-CN"],
        en=idx["en-GB"],
        sk=idx["sk-SK"],
    )


def _protect_placeholders(text: str) -> Tuple[str, List[str]]:
    placeholders: List[str] = []

    def repl(match: re.Match[str]) -> str:
        placeholders.append(match.group(0))
        return f"__PH{len(placeholders) - 1}__"

    return _PLACEHOLDER_RE.sub(repl, text), placeholders


def _restore_placeholders(text: str, placeholders: Sequence[str]) -> str:
    for i, ph in enumerate(placeholders):
        text = text.replace(f"__PH{i}__", ph)
    return text


def _split_ws(text: str) -> Tuple[str, str, str]:
    prefix = re.match(r"^\s*", text, flags=re.DOTALL).group(0)  # type: ignore[union-attr]
    suffix = re.search(r"\s*$", text, flags=re.DOTALL).group(0)  # type: ignore[union-attr]
    core = text[len(prefix) : len(text) - len(suffix)]
    return prefix, core, suffix


def _translate_with_cache(
    texts: Sequence[str],
    translate_batch,
    *,
    cache: Dict[str, str],
) -> List[str]:
    out: List[str] = []
    missing: List[str] = []
    for t in texts:
        if t in cache:
            out.append(cache[t])
        else:
            out.append("")
            missing.append(t)

    if missing:
        translated_missing = translate_batch(missing)
        for src, dst in zip(missing, translated_missing, strict=True):
            cache[src] = dst

        # Fill output in original order
        for i, t in enumerate(texts):
            if out[i] == "":
                out[i] = cache[t]
    return out


def _build_translator(model_name: str):
    from transformers import MarianMTModel, MarianTokenizer
    import torch

    tokenizer = MarianTokenizer.from_pretrained(model_name)
    model = MarianMTModel.from_pretrained(model_name)
    model.eval()

    device = torch.device("cpu")
    model.to(device)

    def translate_batch(texts: Sequence[str]) -> List[str]:
        # Keep deterministic-ish behavior.
        batch_size = 16
        outputs: List[str] = []
        for start in range(0, len(texts), batch_size):
            batch = list(texts[start : start + batch_size])
            encoded = tokenizer(
                batch,
                return_tensors="pt",
                padding=True,
                truncation=True,
                max_length=512,
            )
            encoded = {k: v.to(device) for k, v in encoded.items()}
            generated = model.generate(
                **encoded,
                max_length=512,
                num_beams=4,
            )
            decoded = tokenizer.batch_decode(generated, skip_special_tokens=True)
            outputs.extend(decoded)
        return outputs

    return translate_batch


def _translate_text(
    text: str,
    translate_one,
) -> str:
    prefix, core, suffix = _split_ws(text)
    if not core:
        return text

    protected, placeholders = _protect_placeholders(core)
    # If nothing remains besides placeholders, don't translate.
    if _classify_text(re.sub(r"__PH\d+__", "", protected)) in ("empty", "other"):
        return text

    translated = translate_one(protected)
    restored = _restore_placeholders(translated, placeholders)
    return f"{prefix}{restored}{suffix}"


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Translate TIA/WinCC text export XLSX: Chinese -> English + Slovak; "
            "English -> Slovak. Only fills empty target cells."
        )
    )
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    if not args.input.exists():
        raise FileNotFoundError(str(args.input))
    if args.output.exists():
        raise FileExistsError(f"Refusing to overwrite existing file: {args.output}")

    # Pass 1: collect unique strings that need translation.
    wb_ro = openpyxl.load_workbook(args.input, read_only=True, data_only=True)

    zh_needed: Set[str] = set()
    en_needed_for_sk: Set[str] = set()
    stats = Counter()

    for sheet_name in wb_ro.sheetnames:
        ws = wb_ro[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        cols = _find_columns(headers)

        for row in ws.iter_rows(min_row=2, values_only=True):
            zh_val = row[cols.zh_star] or row[cols.zh]
            en_val = row[cols.en]
            sk_val = row[cols.sk]

            if not _is_blank(zh_val) and _classify_text(str(zh_val)) == "cn":
                if _is_blank(en_val) or _is_blank(sk_val):
                    zh_needed.add(str(zh_val))
                if _is_blank(sk_val) and not _is_blank(en_val) and _classify_text(str(en_val)) == "latin":
                    en_needed_for_sk.add(str(en_val))
            elif not _is_blank(en_val) and _classify_text(str(en_val)) == "latin":
                if _is_blank(sk_val):
                    en_needed_for_sk.add(str(en_val))

    stats["unique_zh_needed"] = len(zh_needed)
    stats["unique_en_needed_for_sk_initial"] = len(en_needed_for_sk)

    # Pass 2: translate.
    print(f"Unique zh->en to translate: {stats['unique_zh_needed']}")
    print(f"Unique en->sk to translate (initial): {stats['unique_en_needed_for_sk_initial']}")

    zh_to_en_batch = _build_translator("Helsinki-NLP/opus-mt-zh-en")
    en_to_sk_batch = _build_translator("Helsinki-NLP/opus-mt-en-sk")

    zh_en_cache: Dict[str, str] = {}
    en_sk_cache: Dict[str, str] = {}

    zh_list = sorted(zh_needed)

    def zh_to_en_one(protected: str) -> str:
        return _translate_with_cache([protected], zh_to_en_batch, cache=zh_en_cache)[0]

    # Translate Chinese -> English on the full source string (placeholders protected).
    for i, zh_src in enumerate(zh_list, start=1):
        zh_en_cache[zh_src] = _translate_text(zh_src, zh_to_en_one)
        if i % 250 == 0:
            print(f"zh->en: {i}/{len(zh_list)}")

    # Expand en_needed_for_sk with derived EN from Chinese for rows missing Slovak.
    for zh_src in zh_list:
        en_needed_for_sk.add(zh_en_cache[zh_src])

    stats["unique_en_needed_for_sk_total"] = len(en_needed_for_sk)
    print(f"Unique en->sk to translate (total, incl derived): {stats['unique_en_needed_for_sk_total']}")

    en_list = sorted(en_needed_for_sk)

    def en_to_sk_one(protected: str) -> str:
        return _translate_with_cache([protected], en_to_sk_batch, cache=en_sk_cache)[0]

    for i, en_src in enumerate(en_list, start=1):
        en_sk_cache[en_src] = _translate_text(en_src, en_to_sk_one)
        if i % 250 == 0:
            print(f"en->sk: {i}/{len(en_list)}")

    # Pass 3: write output by filling only empty target cells.
    wb = openpyxl.load_workbook(args.input)

    fill_stats = Counter()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        cols = _find_columns(headers)

        for r in range(2, ws.max_row + 1):
            zh_val = ws.cell(r, cols.zh_star + 1).value or ws.cell(r, cols.zh + 1).value
            en_cell = ws.cell(r, cols.en + 1)
            sk_cell = ws.cell(r, cols.sk + 1)

            zh_src = str(zh_val) if not _is_blank(zh_val) else ""
            en_src = str(en_cell.value) if not _is_blank(en_cell.value) else ""

            if zh_src and _classify_text(zh_src) == "cn":
                if _is_blank(en_cell.value):
                    en_cell.value = zh_en_cache.get(zh_src, "")
                    if not _is_blank(en_cell.value):
                        fill_stats[f"{sheet_name}:filled_en_from_zh"] += 1
                if _is_blank(sk_cell.value):
                    en_for_sk = en_src or zh_en_cache.get(zh_src, "")
                    sk_cell.value = en_sk_cache.get(en_for_sk, "")
                    if not _is_blank(sk_cell.value):
                        fill_stats[f"{sheet_name}:filled_sk"] += 1
            elif en_src and _classify_text(en_src) == "latin":
                if _is_blank(sk_cell.value):
                    sk_cell.value = en_sk_cache.get(en_src, "")
                    if not _is_blank(sk_cell.value):
                        fill_stats[f"{sheet_name}:filled_sk"] += 1

    wb.save(args.output)

    print("Done.")
    for k, v in sorted(fill_stats.items()):
        print(f"{k} = {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

